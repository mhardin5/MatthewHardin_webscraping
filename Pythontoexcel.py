import openpyxl as xl
from openpyxl.styles import Font


wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1, title='Second Sheet')

ws['A1'] = 'Invoice'
ws['A1'].font = Font(name='Times New Roman', size=24, bold=True)

fontobject = Font(name='Times New Roman', size=24, bold=True)

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

#merge cells
ws.merge_cells('A1:B1')


ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = fontobject

ws['B8'] = '=SUM(B2:B7)'

ws.column_dimensions['A'].width = 15

ws2 = wb['Second Sheet']
read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb('ProduceReport')

ws2['A1'] = 'Produce!'
ws2['B1'] = 'Cost Per Pound'
ws2['C1'] = 'Amt Sold'
ws2['D1'] = 'Total'


row_counter = 2

for row in read_ws.iter_rows(min_row=2):
    name = row[0].value
    cost = row[1].value
    amt_sold = row[2].value
    total = row[3].value

    ws2.cell(row_counter,1) = name
    ws2.cell(row_counter,2) = cost
    ws2.cell(row_counter,3) = amt_sold
    ws2.cell(row_counter,4) = total
    
    row_counter += 1

summary_row1 = row_counter + 1
ws2['B' + str(summary_row1)] = 'Total'
ws2['B' + str(summary_row1)].font = Font(size=16, bold=True)

ws2['C' + str(summary_row1)] = 'SUM(C2:C' + str(row_counter) + ')'
ws2['D' + str(summary_row1)] = 'SUM(D2:D' + str(row_counter) + ')'

summary_row2 = row_counter + 2
ws2['B' + str(summary_row2)] = 'Average'
ws2['B' + str(summary_row2)].font = Font(size=16, bold=True)

ws2['C' + str(summary_row2)] = 'AVERAGE(C2:C' + str(row_counter) + ')'
ws2['D' + str(summary_row2)] = 'AVERAGE(D2:D' + str(row_counter) + ')'

ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 16



wb.save('PythontoExcel.xlsx')

