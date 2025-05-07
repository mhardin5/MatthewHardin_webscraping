
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2025/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
##
##
##

movie_rows = soup.findAll('tr')
wb= xl.Workbook()

ws = wb.active

ws.title = 'Box Office Report'

ws['A1'] = 'Rank'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Release Date'
ws['D1'] = 'Gross Earnings'
ws['E1'] = 'Theaters'
ws['F1'] = 'Avg Gross/Theater'



movie_rows = soup.findAll('tr')
print(movie_rows[1])
input()
for x in range(1,6):
    td = movie_rows[x].findAll('td')
    rank = td[0].text
    title = td[1].text
    theater = int(td[6].text.replace(',', ''))
    total_gross = int(td[5].text.strip('$').replace(",", ""))
    release_date = td[8].text

    avg_per_theater = round(total_gross/theater, 2)
    ws['A' + str(x+1)] = rank
    ws['B' + str(x+1)] = title
    ws['C' + str(x+1)] = release_date
    ws['D' + str(x+1)] = total_gross
    ws['E' + str(x+1)] = theater
    ws['F' + str(x+1)] = avg_per_theater


ws.columns_dimensions['A'].width = 5
ws.columns_dimensions['B'].width = 30
ws.columns_dimensions['C'].width = 25
ws.columns_dimensions['D'].width = 16
ws.columns_dimensions['E'].width = 20
ws.columns_dimensions['F'].width = 26

header_font = Font(size = 16, bold=True)

for cell in ws[1:1]:
    cell.font = header_font

for cell in ws['E:E']:
    cell.number_format = '#,##0'

for cell in ws['D:D']:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws['F:F']:
    cell.number_format = u'"$ "#,##0.00'

wb.save('BoxOfficeReport.xlsx')

